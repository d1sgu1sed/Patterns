import re
from Src.Core.abstract_response import abstract_response
from Src.Core.validator import validator
from Src.Models.recipe_model import recipe_model
from Src.Models.recipe_step_model import recipe_step_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

"""
Исключение при создании Excel файла
"""  
class excel_creation_error(Exception):
    pass


"""
Исключение при форматировании пустой строки
"""  
class empty_description_exception(Exception):
    pass

"""
Исключение при неравном количестве 
плейсхолдеров и параметров шага
"""  
class non_equal_params(Exception):
    pass

"""
Исключение при форматировании
"""  
class formatting_error(Exception):
    pass

class response_xlsx(abstract_response):
    """
    Генератор рецепта в формате Excel (XLSX)

    Поля:
        filename (str): Название файла для загрузки конфига.
        recipe(recipe_model): Модель рецепта.
    """

    __filename: str
    __recipe: recipe_model

    def __init__(self, filename: str, recipe: recipe_model):
        validator.validate(filename, str)
        validator.validate(recipe, recipe_model)
        self.__filename = filename
        self.__recipe = recipe

    """
    Название файла
    """
    @property
    def filename(self):
        return self.__filename
    
    """
    Модель рецепта
    """
    @property
    def recipe(self):
        return self.__recipe

    @filename.setter
    def filename(self, value: str):
        validator.validate(value, str)
        self.__filename = value
    
    @recipe.setter
    def recipe(self, value: recipe_model):
        validator.validate(value, recipe_model)
        self.__recipe = value
    
    """
    Фабричный метод создания
    """
    @staticmethod
    def create(filename: str, recipe: recipe_model):
        validator.validate(filename, str)
        validator.validate(recipe, recipe_model)
        item = response_xlsx(filename, recipe)
        return item

    """
    Форматирует строку description, подставляя параметры из params
    в плейсхолдеры в фигурных скобках.
    """
    def __formatting(self, step: recipe_step_model) -> str:  
        if not step.description:
            raise empty_description_exception("Пустое описание")
        
        if step.params is None:
            step.params = []
        
        placeholders = re.findall(r'\{([^}]+)\}', step.description)
        
        if len(placeholders) != len(step.params):
            raise non_equal_params(
                f"Количество плейсхолдеров {len(placeholders)} не соответствует количеству параметров {len(step.params)}"
            )
        
        try:
            if isinstance(step.params, dict):
                # Для dict-параметров
                formatted_description = step.description.format(**step.params)
            else:
                # Для list-параметров
                formatted_description = step.description.format(*step.params)
            return formatted_description
        except:
            raise formatting_error(
                f"Ошибка при форматировании строки. Плейсхолдеры: {placeholders}, Параметры: {step.params}"
            )

    """
    Форматирует шаги приготовления
    """
    def __format_steps(self) -> list:
        if not self.__recipe.steps:
            return []
        
        formatted_steps = []
        for i, step in enumerate(self.__recipe.steps, 1):
            try:
                formatted_step = self.__formatting(step)
                formatted_steps.append([i, formatted_step])
            except Exception as e:
                formatted_steps.append([i, f"[Ошибка форматирования шага: {e}]"])
        
        return formatted_steps

    """
    Форматирует таблицу ингредиентов
    """
    def __format_ingredients_table(self) -> list:
        if not self.__recipe.ingredients:
            return []
        
        ingredients_data = []
        for ingredient in self.__recipe.ingredients:
            name = ingredient.product.name
            quantity = f"{ingredient.amount} {ingredient.product.measure.name}"
            ingredients_data.append([name, quantity])
        
        return ingredients_data

    """
    Генерирует Excel файл с рецептом
    """
    def generate(self) -> str:
        try:
            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Рецепт"
            
            # Стили
            header_font = Font(bold=True, size=14)
            title_font = Font(bold=True, size=12)
            normal_font = Font(size=11)
            center_align = Alignment(horizontal='center')
            
            current_row = 1
            
            # Заголовок рецепта
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'] = self.__recipe.name.upper()
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].alignment = center_align
            current_row += 2
            
            # Таблица ингредиентов
            if self.__recipe.ingredients:
                ws[f'A{current_row}'] = "ИНГРЕДИЕНТЫ"
                ws[f'A{current_row}'].font = title_font
                current_row += 1
                
                # Заголовки таблицы ингредиентов
                ws[f'A{current_row}'] = "Ингредиенты"
                ws[f'B{current_row}'] = "Количество"
                for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
                    cell.font = title_font
                current_row += 1
                
                # Данные ингредиентов
                ingredients_data = self.__format_ingredients_table()
                for ingredient in ingredients_data:
                    ws[f'A{current_row}'] = ingredient[0]
                    ws[f'B{current_row}'] = ingredient[1]
                    for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
                        cell.font = normal_font
                    current_row += 1
                
                current_row += 1
            
            # Ремарка
            if self.__recipe.remark:
                ws[f'A{current_row}'] = "РЕМАРКА"
                ws[f'A{current_row}'].font = title_font
                current_row += 1
                ws[f'A{current_row}'] = self.__recipe.remark
                ws[f'A{current_row}'].font = normal_font
                current_row += 2
            
            # Шаги приготовления
            if self.__recipe.steps:
                ws[f'A{current_row}'] = "ПОШАГОВОЕ ПРИГОТОВЛЕНИЕ"
                ws[f'A{current_row}'].font = title_font
                current_row += 1
                
                # Заголовки таблицы шагов
                ws[f'A{current_row}'] = "Шаг"
                ws[f'B{current_row}'] = "Описание"
                for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
                    cell.font = title_font
                current_row += 1
                
                # Данные шагов
                steps_data = self.__format_steps()
                for step in steps_data:
                    ws[f'A{current_row}'] = step[0]
                    ws[f'B{current_row}'] = step[1]
                    for cell in [ws[f'A{current_row}'], ws[f'B{current_row}']]:
                        cell.font = normal_font
                    # Перенос текста для описания
                    ws[f'B{current_row}'].alignment = Alignment(wrap_text=True)
                    current_row += 1
            
            # Настраиваем ширину колонок
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 80
            
            # Сохраняем файл
            wb.save(self.__filename)
            
            return f"Excel файл успешно создан: {self.__filename}"
            
        except Exception as e:
            raise excel_creation_error(f"Ошибка при создании Excel файла: {e}")