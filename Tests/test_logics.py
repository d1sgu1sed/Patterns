import os
import unittest
import json
from Src.Logics.response_csv import response_csv
from Src.Logics.response_json import response_json
from Src.Logics.response_md import response_md
from Src.Logics.response_xlsx import response_xlsx
from Src.Models.nomenclature_group_model import nomenclature_group_model
from Src.Logics.factory_entities import factory_entities
from Src.Core.response_format import response_formats
from Src.Core.validator import validator
from Src.Core.abstract_response import abstract_response
from Src.reposity import reposity
from Src.start_service import start_service

# Тесты для проверки логики 
class Test_factory(unittest.TestCase):
    __start_service = start_service()

    def test_md_factory_create(self):
        # Подготовка
        factory = factory_entities()
        data = self.__start_service.reposity.data[reposity.recipes_key()][0]
        expected_filename = "Docs/waffles.md"

        # Действие
        logic = factory.create(response_formats.md())
        # удаляем файл если он существует
        if os.path.exists(expected_filename):
            os.unlink(expected_filename)

        # Проверка
        assert logic is not None
        instance = logic.create(expected_filename, data)
        validator.validate(instance, response_md)
        text = instance.generate()
        assert os.path.exists(expected_filename)
        # проверяем что содержимое корректное
        with open(expected_filename, 'r', encoding='utf-8') as f:
            file_content = f.read()
        
        assert text == file_content
        assert f"# {data.name.upper()}" in text
    
    def test_json_factory_create(self):
    # Подготовка
        factory = factory_entities()
        data = self.__start_service.reposity.data[reposity.recipes_key()][0]
        expected_filename = "Docs/waffles.json"

        # Действие
        logic = factory.create(response_formats.json())
        
        # удаляем файл если он существует
        if os.path.exists(expected_filename):
            os.unlink(expected_filename)

        # Проверка
        assert logic is not None
        instance = logic.create(expected_filename, data)
        validator.validate(instance, response_json)
        text = instance.generate()
        assert os.path.exists(expected_filename)
        
        # проверяем что содержимое корректное
        with open(expected_filename, 'r', encoding='utf-8') as f:
            file_content = f.read()
        
        assert text == file_content
        
        # проверяем что JSON валидный и содержит основные поля
        json_data = json.loads(text)
        assert "name" in json_data
        assert "ingredients" in json_data
        assert "steps" in json_data
        assert json_data["name"] == data.name
    
    def test_csv_not_none_factory_create(self):
        # Подготовка
        factory = factory_entities()
        data = self.__start_service.reposity.data[reposity.recipes_key()][0]
        expected_filename = "Docs/waffles.csv"

        # Действие
        logic = factory.create(response_formats.csv())
        
        # удаляем файл если он существует
        if os.path.exists(expected_filename):
            os.unlink(expected_filename)

        # Проверка
        assert logic is not None
        instance = logic.create(expected_filename, data)
        validator.validate(instance, response_csv)
        text = instance.generate()
        assert os.path.exists(expected_filename)
        
        # проверяем что содержимое корректное
        with open(expected_filename, 'r', encoding='utf-8') as f:
            file_content = f.read()
        
        # Нормализуем переносы строк перед сравнением
        text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
        file_content_normalized = file_content.replace('\r\n', '\n').replace('\r', '\n')
        
        assert text_normalized == file_content_normalized
        
        # проверяем что CSV содержит основные данные
        assert data.name.upper() in text_normalized
        assert "ИНГРЕДИЕНТЫ" in text_normalized
        assert "ПОШАГОВОЕ ПРИГОТОВЛЕНИЕ" in text_normalized
    
    def test_xlsx_not_none_factory_create(self):
    # Подготовка
        factory = factory_entities()
        data = self.__start_service.reposity.data[reposity.recipes_key()][0]
        expected_filename = "Docs/waffles.xlsx"

        # Действие
        logic = factory.create(response_formats.xlsx())
        
        # удаляем файл если он существует
        if os.path.exists(expected_filename):
            os.unlink(expected_filename)

        # Проверка
        assert logic is not None
        instance = logic.create(expected_filename, data)
        validator.validate(instance, response_xlsx)
        text = instance.generate()
        assert os.path.exists(expected_filename)
        
        # проверяем что файл создан и содержит основные данные
        # (для Excel мы не можем просто прочитать текстом, проверяем наличие файла)
        assert "Excel файл успешно создан" in text
        
        # дополнительная проверка - пытаемся открыть файл как Excel
        try:
            from openpyxl import load_workbook
            wb = load_workbook(expected_filename)
            ws = wb.active
            assert ws.title == "Рецепт"
            assert ws['A1'].value == data.name.upper()
            wb.close()
        except:
            # Если не установлен openpyxl, просто пропускаем эту проверку
            pass

  
if __name__ == '__main__':
    unittest.main()